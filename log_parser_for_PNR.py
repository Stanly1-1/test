import os
import re
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# -------------------------
# 🔧 Парсинг отдельных данных из логов
# -------------------------

#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
# Блок в функциями 

def extract_raid_info(text): #ПОДУМАТЬ ЗА 2йо рейд в сервере, а если их будет 3? 
    """
    Возвращает список всех строк RAID вида: 0/239 RAID1 ... из блока 'storcli64 /c0/vall show'
    """
    lines = text.splitlines()
    raid_entries = []
    in_block = False

    for line in lines:
        if re.match(r">>> storcli64 /c0/vall show:", line):
            in_block = True
        elif in_block and re.match(r"^-{10,}$", line):  # первая разделительная линия
            continue
        elif in_block and re.match(r'^\d+/\d+\s+RAID\d\s', line.strip()):
            raid_entries.append(line.strip())
        elif in_block and "VD=Virtual Drive" in line:
            break  # конец таблицы
    return raid_entries


def extract_bios_date(text):
    """
    Извлекает строку даты BIOS из блока >>> date BIOS:
    """
    match = re.search(r'>>> date BIOS:\s*\n([^\n]+)', text)
    return match.group(1).strip() if match else ""







def extract_sn_and_product_name_(text):
    """
    Ищет блок platform (SRV) и парсит оттуда Serial Number и Product Name.
    Добавлен отладочный вывод.
    Учитывает строки с разделителем '│' и без него (fixed-width parsing).
    """
    print(">>> 🔍 Ищу блок 'platform (SRV)'")
    platform_match = re.search(
        r">>> sds-inventory-manager platform \(SRV\):\n(.*?)(>>>|\Z)",
        text,
        re.DOTALL
    )

    if not platform_match:
        print("❌ Блок не найден.")
        return "", ""

    block = platform_match.group(1)
    print("✅ Найденный блок:\n" + "-" * 40)
    print(block)
    print("-" * 40)

    for line in block.splitlines():
        print(f"🔹 Строка: {line}")
        if "VEGMAN" in line.upper():
            print("✅ Найдена строка с VEGMAN")

            if "│" in line:
                # Разделение по символу '│' с очисткой пробелов
                clean_line = re.sub(r'[│|]', '|', line)
                parts = [p.strip() for p in clean_line.split('|') if p.strip()]
                print(f"🔧 Разделено на части (по '│'): {parts}")

                if len(parts) >= 3:
                    product_name = parts[1]
                    serial_number = parts[2]
                    print(f"✅ SN: {serial_number}, Product: {product_name}")
                    return serial_number, product_name
                else:
                    print("❌ Недостаточно данных в строке с разделителем '│'")
            else:
                # Парсим по фиксированным позициям (fixed-width parsing)
                # Примерные позиции — можно подкорректировать под формат файла
                manufacturer = line[0:13].strip()
                product_name = line[13:38].strip()
                serial_number = line[38:52].strip()
                sku_number = line[52:].strip()
                parts = [manufacturer, product_name, serial_number, sku_number]
                print(f"🔧 Разделено на части (fixed-width): {parts}")

                if product_name and serial_number:
                    print(f"✅ SN: {serial_number}, Product: {product_name}")
                    return serial_number, product_name
                else:
                    print("❌ Недостаточно данных в строке fixed-width")

        else:
            print("⏭️ Строка не содержит VEGMAN")

    print("❌ Строка с VEGMAN не найдена в блоке")
    return "", ""





def extract_firmware_versions(text):
    bmc_version = bios_version = fpga_version = ""
    lines = text.splitlines()
    for idx, line in enumerate(lines):
        if 'bmc info version' in line:
            for l in lines[idx + 1:idx + 6]:
                if l.strip().startswith("Host") and not bios_version:
                    parts = l.strip().split(maxsplit=1)
                    bios_version = parts[1] if len(parts) == 2 else ""
                elif l.strip().startswith("BMC") and not bmc_version:
                    parts = l.strip().split(maxsplit=1)
                    bmc_version = parts[1] if len(parts) == 2 else ""
            break
    fpga_match = re.search(r'FPGA firmware version>>>\s*([^\s"\']+)', text)
    if fpga_match:
        fpga_version = fpga_match.group(1)
    return {"bmc": bmc_version, "bios": bios_version, "fpga": fpga_version}


def extract_json_after_command(text, command_marker):
    lines = text.splitlines()
    json_lines = []
    json_started = False
    brace_balance = 0

    for idx, line in enumerate(lines):
        if command_marker in line:
            # начинаем проверку после найденной команды
            for subline in lines[idx+1:]:
                subline_stripped = subline.strip()

                if not json_started:
                    if subline_stripped.startswith('{'):
                        json_started = True
                    else:
                        continue

                if json_started:
                    json_lines.append(subline)
                    brace_balance += subline.count('{')
                    brace_balance -= subline.count('}')
                    if brace_balance == 0:
                        break
            break

    if not json_lines:
        return None

    json_str = "\n".join(json_lines)
    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        print(f"[WARN] Ошибка разбора JSON: {e}")
        return None


def extract_system_info(text):
    cpu = re.search(r'Physical CPU Count:\s*(\d+)', text)
    ram = re.search(r'Total RAM:\s*(\d+\s*GB)', text)
    return {
        "cpu_count": cpu.group(1) if cpu else "",
        "total_ram": ram.group(1) if ram else ""
    }

def extract_health_status(text):
    lines = text.splitlines()
    for idx, line in enumerate(lines):
        if re.search(r'health sensors\s*\|\s*grep\s+-E\s+[\'"]?Warning\|Critical[\'"]?', line):
            for j in range(idx + 1, min(idx + 6, len(lines))):
                next_line = lines[j].strip()
                if not next_line:
                    continue
                if re.match(r'^\w+@[\w\-]+:.*\$', next_line):
                    return "✅"
                return "❌"
            return "✅"
    return "❌"

def extract_p3v3_value(text):
    for line in text.splitlines():
        if re.search(r'\bP3V3\b', line):
            match = re.search(r'P3V3\s+\w+\s+([\d.]+)\s+V', line)
            if match:
                return match.group(1)
    return ""

def extract_sdcard_status(text):
    return "✅" if "/dev/mmcblk0p1" in text else "❌"

def format_items(items, kind="storage"):
    result = []
    for item in items:
        pn = item.get('product_name', 'N/A')
        fw = item.get('firmware_version', 'N/A')
        psoc = item.get('psoc_firmware_version', 'N/A') if kind == "storage" else None
        result.append(f"{pn} ({fw} / {psoc})" if kind == "storage" else f"{pn} ({fw})")
    return result
    
def extract_ifconfig_block(text):
    """
    Извлекает параметры из блока 'bmc ifconfig show :'
    """
    result = {
        "host_name": "",
        "ipv4_gateway": "",
        "ipv6_gateway": "",
        "eth0": {
            "ip": "",
            "dhcp": "",
            "dns": "",
            "static_dns": "",
            "ntp": ""
        },
        "eth1": {
            "ip_list": [],
            "dhcp": "",
            "dns": "",
            "static_dns": "",
            "ntp": ""
        }
    }

    lines = text.splitlines()
    in_block = False
    current_section = None

    for i, line in enumerate(lines):
        if "bmc ifconfig show :" in line:
            in_block = True
            continue
        if not in_block:
            continue

        line = line.strip()

        if not line:
            continue

        if line.startswith("Global network configuration"):
            current_section = "global"
        elif line.startswith("Management ethernet interface (eth0):"):
            current_section = "eth0"
        elif line.startswith("Switched ethernet interface (eth1):"):
            current_section = "eth1"
        elif line.startswith("Ethernet interface (sit0):"):
            break  # конец нужного блока

        elif current_section == "global":
            if "Host name:" in line:
                result["host_name"] = line.split(":", 1)[1].strip()
            elif "Default IPv4 gateway:" in line:
                result["ipv4_gateway"] = line.split(":", 1)[1].strip()
            elif "Default IPv6 gateway:" in line:
                result["ipv6_gateway"] = line.split(":", 1)[1].strip()

        elif current_section == "eth0":
            if "IP address:" in line:
                result["eth0"]["ip"] = line.split(":", 1)[1].strip()
            elif "DHCP:" in line:
                result["eth0"]["dhcp"] = line.split(":", 1)[1].strip()
            elif "DNS servers:" in line:
                result["eth0"]["dns"] = line.split(":", 1)[1].strip()
            elif "Static DNS servers:" in line:
                result["eth0"]["static_dns"] = line.split(":", 1)[1].strip()
            elif "NTP servers:" in line:
                result["eth0"]["ntp"] = line.split(":", 1)[1].strip()

        elif current_section == "eth1":
            if "IP address:" in line:
                result["eth1"]["ip_list"].append(line.split(":", 1)[1].strip())
            elif "DHCP:" in line:
                result["eth1"]["dhcp"] = line.split(":", 1)[1].strip()
            elif "DNS servers:" in line:
                result["eth1"]["dns"] = line.split(":", 1)[1].strip()
            elif "Static DNS servers:" in line:
                result["eth1"]["static_dns"] = line.split(":", 1)[1].strip()
            elif "NTP servers:" in line:
                result["eth1"]["ntp"] = line.split(":", 1)[1].strip()

    return result


# -------------------------
# 📂 Обработка лог-файла   !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
# -------------------------

def process_file(filepath, debug=False):
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        text = f.read()

    sn, product_name = extract_sn_and_product_name_(text)
    json_data = extract_json_after_command(text, '>>> Start check sds-inventory-manager get (PCI):')
    fwinfo = extract_firmware_versions(text)
    sysinfo = extract_system_info(text)
    health = extract_health_status(text)
    p3v3_value = extract_p3v3_value(text)
    mgmt_info = extract_ifconfig_block(text)
    sdcard_status = "✅" if "/dev/mmcblk0p1" in text else "❌"
    bios_date = extract_bios_date(text)
    raid_info_list = extract_raid_info(text)
    
    
    storage = format_items(json_data.get('storage_controllers', []), 'storage') if json_data else []
    fibre = format_items(json_data.get('fibre_channel_adapters', []), 'fibre') if json_data else []
    network = format_items(json_data.get('network_adapters', []), 'network') if json_data else []
    
    
    disks = []
    if json_data:
        for d in json_data.get("disk_drives", []):
            disks.append({
                "manufacturer": d.get("manufacturer", ""),
                "product_name": d.get("product_name", ""),
                "firmware_version": d.get("firmware_version", "")
            })
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#DEBUG 

    if debug:
        print(f"\n===> Файл: {filepath}")
        print(f"[{'OK' if sn else 'FAIL'}] SN: {sn or 'не найден'}")
        print(f"[{'OK' if json_data else 'FAIL'}] JSON блок {'найден' if json_data else 'не найден'}")
        print(f"[{'OK' if all([fwinfo.get('bmc'), fwinfo.get('bios'), fwinfo.get('fpga')]) else 'FAIL'}] Прошивки: BMC/BIOS/FPGA")
        print(f"[{'OK' if sysinfo.get('cpu_count') and sysinfo.get('total_ram') else 'FAIL'}] CPU и RAM")
        print(f"[{'OK' if health.strip() == '✅' else 'FAIL'}] Health status")
        print(f"[{'OK' if p3v3_value else 'FAIL'}] P3V3: {p3v3_value or 'нет'}")
        print(f"[{'OK' if sdcard_status == '✅' else 'FAIL'}] SDcard")
        print(f"[{'OK' if raid_info_list else 'FAIL'}] RAID entries: {len(raid_info_list)}")
        
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    else:
        ok = all([
            sn, json_data,
            fwinfo.get("bmc"), fwinfo.get("bios"), fwinfo.get("fpga"),
            sysinfo.get("cpu_count"), sysinfo.get("total_ram")
        ])
        status = "[OK]" if ok else "[FALSE]"
        message = "успешно обработан" if ok else "неполные данные"
        print(f"{status} SN {sn or 'UNKNOWN'} ({os.path.basename(filepath)}) — {message}")


    return {
        "sn": sn,
        "cpu": sysinfo.get("cpu_count", ""),
        "ram": sysinfo.get("total_ram", ""),
        "storage": storage,
        "fibre": fibre,
        "network": network,
        "bmc": fwinfo.get("bmc", ""),
        "bios": fwinfo.get("bios", ""),
        "fpga": fwinfo.get("fpga", ""),
        "health": health.strip(),
        "p3v3": p3v3_value,
        "sdcard": sdcard_status,
        "disks": disks,
        "product_name": product_name,
        "bios_date": bios_date,
        "raid_info": raid_info_list,
        "mgmt": extract_ifconfig_block(text)
        
        
    }


# -------------------------
# 📊 Генерация Excel
# -------------------------

def save_to_excel(results):
    wb = Workbook()

    def create_sheet(title, headers, rows):
        ws = wb.create_sheet(title)
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="0000CC")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 12)
            
            
            
#-----------------------------------------------------------------------------
    # 🧾 Summary
    create_sheet(
        "Summary",
        ["SN", "Product Name", "BIOS Date", "Health Status", "P3V3 (V)", "SDcard", "BMC", "BIOS", "FPGA", "CPU", "RAM"],
        [[
            r["sn"],
            r.get("product_name", ""),
            r.get("bios_date", ""),
            r.get("health", ""),
            r.get("p3v3") or "N/A",
            r.get("sdcard", "❌"),
            r["bmc"],
            r["bios"],
            r["fpga"],
            r["cpu"],
            r["ram"]
        ] for r in results]
    )
    

    max_storage = max(len(r["storage"]) for r in results)
    create_sheet("Storage", ["SN"] + [f"Storage #{i+1}" for i in range(max_storage)],
        [[r["sn"]] + r["storage"] + [""] * (max_storage - len(r["storage"])) for r in results])

    max_fc = max(len(r["fibre"]) for r in results)
    create_sheet("FC", ["SN"] + [f"FC #{i+1}" for i in range(max_fc)],
        [[r["sn"]] + r["fibre"] + [""] * (max_fc - len(r["fibre"])) for r in results])

    max_net = max(len(r["network"]) for r in results)
    create_sheet("Network", ["SN"] + [f"Network #{i+1}" for i in range(max_net)],
        [[r["sn"]] + r["network"] + [""] * (max_net - len(r["network"])) for r in results])

    disk_headers = ["SN", "Disk Count", "Manufacturer", "Product Name", "Firmware Version"]
    disk_rows = []
    for r in results:
        sn = r["sn"]
        disks = r.get("disks", [])
        disk_count = len(disks)
        if not disks:
            disk_rows.append([sn, 0, "", "", ""])
        else:
            for d in disks:
                disk_rows.append([sn, disk_count, d["manufacturer"], d["product_name"], d["firmware_version"]])
    create_sheet("Disks", disk_headers, disk_rows)

#-----------------------------------------------------------------------------

    # 🌐 Лист MGMT (Management Network Info)
    mgmt_headers = ["SN", "Host Name", "IPv4 Gateway", "IPv6 Gateway",
                    "eth0 IP", "eth0 DHCP", "eth0 DNS", "eth0 Static DNS", "eth0 NTP",
                    "eth1 IP #1", "eth1 IP #2", "eth1 IP #3",
                    "eth1 DHCP", "eth1 DNS", "eth1 Static DNS", "eth1 NTP"]

    mgmt_rows = []
    for r in results:
        mgmt = r.get("mgmt", {})
        eth1_ips = mgmt.get("eth1", {}).get("ip_list", [])
        row = [
            r["sn"],
            mgmt.get("host_name", ""),
            mgmt.get("ipv4_gateway", ""),
            mgmt.get("ipv6_gateway", ""),
            mgmt.get("eth0", {}).get("ip", ""),
            mgmt.get("eth0", {}).get("dhcp", ""),
            mgmt.get("eth0", {}).get("dns", ""),
            mgmt.get("eth0", {}).get("static_dns", ""),
            mgmt.get("eth0", {}).get("ntp", ""),
            eth1_ips[0] if len(eth1_ips) > 0 else "",
            eth1_ips[1] if len(eth1_ips) > 1 else "",
            eth1_ips[2] if len(eth1_ips) > 2 else "",
            mgmt.get("eth1", {}).get("dhcp", ""),
            mgmt.get("eth1", {}).get("dns", ""),
            mgmt.get("eth1", {}).get("static_dns", ""),
            mgmt.get("eth1", {}).get("ntp", "")
        ]
        mgmt_rows.append(row)

    create_sheet("MGMT", mgmt_headers, mgmt_rows)
    
    
#-----------------------------------------------------------------------------
    # Считаем максимальное число RAID строк среди всех логов
    max_raid_count = max(len(r.get("raid_info", [])) for r in results)

    # Формируем заголовки RAID #1, RAID #2 и т.д.
    raid_headers = ["SN"] + [f"RAID #{i+1}" for i in range(max_raid_count)] + ["BBU Status"]

    # Формируем строки
    raid_rows = []
    for r in results:
        raids = r.get("raid_info", [])
        raid_row = [r["sn"]] + raids + [""] * (max_raid_count - len(raids)) + [r.get("bbu_status", "")]
        raid_rows.append(raid_row)

    # Создаем лист
    create_sheet("RAID", raid_headers, raid_rows)
    
#-----------------------------------------------------------------------------
    
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
#ПРАВИТЬ ПУТЬ ТУТ!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    wb.save("/mnt/c/tmp/EasyPNR/output.xlsx")
    print("✅ Excel-файл 'output.xlsx' успешно создан.")

# -------------------------
# 🚀 Точка входа
# -------------------------

def print_help():
    print(r"""
                                                                                                                                  
 _                                                     __            ______ _   _ ______ 
| |                                                   / _|           | ___ \ \ | || ___ \
| |     ___   __ _   _ __   __ _ _ __ ___  ___ _ __  | |_ ___  _ __  | |_/ /  \| || |_/ /
| |    / _ \ / _` | | '_ \ / _` | '__/ __|/ _ \ '__| |  _/ _ \| '__| |  __/| . ` ||    / 
| |___| (_) | (_| | | |_) | (_| | |  \__ \  __/ |    | || (_) | |    | |   | |\  || |\ \ 
\_____/\___/ \__, | | .__/ \__,_|_|  |___/\___|_|    |_| \___/|_|    \_|   \_| \_/\_| \_|
              __/ | | |                                                                  
             |___/  |_|                                                                  

Log parser for PNR — инструмент для разбора логов серверов при ПНР.

Использование:
  python3 log_parser_for_PNR.py [ --debug ] [ --help | -h ]

Параметры:
  --debug     Включает подробный режим отладки
  --help      Показывает это справочное сообщение
 
Описиние:
- Скрипт для парсинга логов серверов при ПНР и генерации отчёта в Excel.
- Автоматически извлекает информацию о SN, health , PCI адаптеры и их прошивки и т.д. 
- Обрабатывает все .log файлы в текущем каталоге.

        __   _____ ____________ _____                                                    
        \ \ / / _ \|  _  \ ___ \  _  |                                                   
         \ V / /_\ \ | | | |_/ / | | |                                                   
          \ /|  _  | | | |    /| | | |                                                   
 _ _ _    | || | | | |/ /| |\ \\ \_/ /  _ _ _                                            
(_|_|_)   \_/\_| |_/___/ \_| \_|\___/  (_|_|_)   
  

""")


def main():
    # Если запрошен help — показать и завершить
    if "--help" in sys.argv or "-h" in sys.argv:
        print_help()
        return

    debug = "--debug" in sys.argv or "-d" in sys.argv
    results = []
#ПРАВИТЬ ПУТЬ ТУТ!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

    log_dir = "/mnt/c/tmp/EasyPNR/merged_logs"  # укажи здесь нужный путь к папке с логами

    for file in os.listdir(log_dir):
        if file.endswith('.log'):
            filepath = os.path.join(log_dir, file)
            results.append(process_file(filepath, debug=debug))


    if results:
        save_to_excel(results)
    else:
        print("⚠️ Нет подходящих логов.")



if __name__ == "__main__":
    main()
